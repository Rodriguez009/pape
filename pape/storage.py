import fcntl
import os
import re
import shutil
import tempfile
from contextlib import contextmanager
from dataclasses import dataclass, asdict
from typing import List, Optional

from openpyxl import Workbook, load_workbook

from .utils import warn


HOME = os.path.expanduser("~")
DATA_DIR = os.path.join(HOME, "pape")
PDF_DIR = os.path.join(DATA_DIR, "pdf")
XLSX_PATH = os.path.join(DATA_DIR, "info.xlsx")
LOCK_PATH = os.path.join(DATA_DIR, ".info.xlsx.lock")

COLUMNS = ["id", "url", "title", "abstract", "submit_date", "path", "added_date"]


@dataclass
class Row:
    id: str = ""
    url: str = ""
    title: str = ""
    abstract: str = ""
    submit_date: str = ""
    path: str = ""
    added_date: str = ""

    def as_list(self):
        return [getattr(self, c) for c in COLUMNS]

    @classmethod
    def from_list(cls, values):
        d = {c: ("" if v is None else str(v)) for c, v in zip(COLUMNS, values)}
        return cls(**d)


def ensure_dirs() -> None:
    os.makedirs(PDF_DIR, exist_ok=True)


def _ensure_xlsx() -> None:
    if os.path.exists(XLSX_PATH):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "papers"
    ws.append(COLUMNS)
    wb.save(XLSX_PATH)


@contextmanager
def _flock():
    """Exclusive file lock so concurrent invocations don't clobber the xlsx."""
    ensure_dirs()
    fd = os.open(LOCK_PATH, os.O_CREAT | os.O_RDWR, 0o644)
    try:
        fcntl.flock(fd, fcntl.LOCK_EX)
        yield
    finally:
        try:
            fcntl.flock(fd, fcntl.LOCK_UN)
        finally:
            os.close(fd)


def load_rows() -> List[Row]:
    ensure_dirs()
    _ensure_xlsx()
    try:
        wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"无法读取 {XLSX_PATH}: {exc}") from exc
    ws = wb.active
    rows: List[Row] = []
    first = True
    for raw in ws.iter_rows(values_only=True):
        if first:
            first = False
            # Skip header row whether or not it matches exactly
            if raw and str(raw[0] or "").lower() == "id":
                continue
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        padded = list(raw) + [None] * (len(COLUMNS) - len(raw))
        rows.append(Row.from_list(padded[: len(COLUMNS)]))
    wb.close()
    return rows


def save_rows(rows: List[Row]) -> None:
    """Atomic save: write to temp file, then replace."""
    ensure_dirs()
    wb = Workbook()
    ws = wb.active
    ws.title = "papers"
    ws.append(COLUMNS)
    for r in rows:
        ws.append(r.as_list())
    tmp_fd, tmp_path = tempfile.mkstemp(
        prefix=".info.", suffix=".xlsx", dir=DATA_DIR
    )
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, XLSX_PATH)
    except PermissionError as exc:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise RuntimeError(
            f"写入 {XLSX_PATH} 被拒绝（{exc}）。如果该文件在 Excel/Numbers 中打开，请先关闭。"
        ) from exc
    except Exception as exc:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise RuntimeError(f"写入 {XLSX_PATH} 失败: {exc}") from exc


@contextmanager
def write_session():
    """Lock + load + yield rows for in-place mutation + save."""
    with _flock():
        rows = load_rows()
        state = {"rows": rows, "save": True}
        yield state
        if state["save"]:
            save_rows(state["rows"])


def _strip_version(s: str) -> str:
    return re.sub(r"v\d+$", "", s or "")


def find_rows(rows: List[Row], needle: str) -> List[int]:
    """Match priority: id exact (ignoring vN) -> title exact (ci) -> title substring (ci).
    Returns indices into rows."""
    if not needle:
        return []
    needle = needle.strip()
    nl = needle.lower()
    n_stripped = _strip_version(needle)

    # 1. id exact
    hits = [i for i, r in enumerate(rows) if _strip_version(r.id) == n_stripped]
    if hits:
        return hits
    # 2. title exact (ci)
    hits = [i for i, r in enumerate(rows) if r.title.lower() == nl]
    if hits:
        return hits
    # 3. title substring (ci)
    hits = [i for i, r in enumerate(rows) if nl in r.title.lower()]
    return hits


def find_by_id_exact(rows: List[Row], arxiv_id: str) -> Optional[int]:
    base = _strip_version(arxiv_id)
    for i, r in enumerate(rows):
        if _strip_version(r.id) == base:
            return i
    return None


def pick_pdf_path(safe_stem: str, arxiv_id: str) -> str:
    """Return a non-conflicting absolute path under PDF_DIR for a new PDF."""
    candidate = os.path.join(PDF_DIR, safe_stem + ".pdf")
    if not os.path.exists(candidate):
        return candidate
    # Append arxiv id to disambiguate
    base = _strip_version(arxiv_id).replace("/", "_")
    candidate = os.path.join(PDF_DIR, f"{safe_stem}__{base}.pdf")
    if not os.path.exists(candidate):
        return candidate
    # Last resort: numeric suffix
    n = 2
    while True:
        candidate = os.path.join(PDF_DIR, f"{safe_stem}__{base}_{n}.pdf")
        if not os.path.exists(candidate):
            return candidate
        n += 1
